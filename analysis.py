import aspen_tools
import pandas as pd
import numpy as np
import time
import pythoncom
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

def calc_Tb(Aspen,Asp, T_list,P_list, NCOMP):
    Tl = T_list[0]
    Tu = T_list[-1]
    Tb_list = []
    Asp.set_focus()
    aspen_tools.cd_properties(Asp)
    Asp.child_window(title="PURE-G", control_type="TreeItem").click_input()
    for Pi, P in enumerate(P_list):
        n = 200
        Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\PURE-G\Input\PLIST\#0").Value = P
        Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\PURE-G\Input\TLOWER").Value = Tl
        Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\PURE-G\Input\TUPPER").Value = Tu
        Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\PURE-G\Input\TNPOINT").Value = n
        aspen_tools.run(Asp)
        time.sleep(2)  
        result = [0] * NCOMP  # 對每個成分記錄沸點或狀態
        for k in range(NCOMP):
            for i in range(n + 1):
                lG = Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\PURE-G\Output\Prop Data\PROPTAB").Elements.Item(f"LIQUID G COMP{k+1}", str(i+1)).Value
                vG = Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\PURE-G\Output\Prop Data\PROPTAB").Elements.Item(f"VAPOR G COMP{k+1}", str(i+1)).Value
                if lG > vG and i == 0:
                    result[k] = "TB is under Tl"
                    break
                elif lG >= vG:
                    result[k] = round(Tl + i * (Tu - Tl) / n, 2)
                    break

            if i == n and result[k] == 0:
                result[k] = "TB is above Tu"
        Tb_list.append(result)
    format_Tb_table(Tb_list, P_list, Tl, Tu)
    return Tb_list

def format_Tb_table(Tb_list, P_list, Tl, Tu):
    comp_aliases = [f"COMP{i+1}" for i in range(len(Tb_list[0]))]

    col_w = 18
    left_w = 14
    total_w = left_w + (len(comp_aliases) * (col_w + 3))

    print("=" * total_w)
    header = f"{'Pressure':<{left_w}}" + "".join([f"| {alias:<{col_w}}" for alias in comp_aliases])
    print(header)
    print("=" * total_w)

    for p, row in zip(P_list, Tb_list):
        line = f"{f'{p} atm':<{left_w}}"
        for v in row:
            if isinstance(v, str):
                if "under Tl" in v:
                    v = f"< {Tl} K"
                elif "above Tu" in v:
                    v = f"> {Tu} K"
            line += f"| {str(v) :<{col_w}} K"
        print(line)

    print("=" * total_w)



def Bpure(Aspen, Asp, T_list, P_list, Tb_list, n_list,NCOMP,O = False):
    Tl = T_list[0]
    Tu = T_list[-1]
    rxnB_list = []
    aspen_tools.cd_properties(Asp)
    # Asp.set_focus()
    # Asp.child_window(title="PURE-B", control_type="TreeItem").click_input()
    index = 0
    Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\PURE-B\Input\TLOWER").Value = Tl
    Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\PURE-B\Input\TUPPER").Value = Tu
    Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\PURE-B\Input\TNPOINT").Value = len(T_list) - 1
    for Pi, P in enumerate(P_list):
        rxnB = []
        Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\PURE-B\Input\PLIST\#0").Value = P
        aspen_tools.run(Asp)

        for n, T in enumerate(T_list):
            B = []
            rB = 0

            for i in range(NCOMP):
                vB = Aspen.Application.Tree.FindNode(
                    r"\Data\Properties\Analysis\PURE-B\Output\Prop Data\PROPTAB"
                ).Elements.Item(f"VAPOR AVAIL COMP{i+1}", str(n + 1)).Value

                lB = Aspen.Application.Tree.FindNode(
                    r"\Data\Properties\Analysis\PURE-B\Output\Prop Data\PROPTAB"
                ).Elements.Item(f"LIQUID AVAIL COMP{i+1}", str(n + 1)).Value

                Tb = Tb_list[Pi][i]
                if Tb == "TB is under Tl":
                    B.append(vB)
                elif Tb == "TB is above Tu":
                    B.append(lB)
                elif T >= Tb:
                    B.append(vB)
                else:
                    B.append(lB)
            if O and n==0 and Pi ==0:#算n_out_Bp0
                B0 = B
            for i in range(NCOMP):
                if isinstance(n_list[0],list):
                    if O:
                        rB += n_list[index][i] * B0[i]
                    else:
                        rB += n_list[index][i] * B[i]
                else:
                    rB += n_list[i] * B[i]

            rxnB.append(round(rB, 7))
            index += 1

        rxnB_list.append(rxnB)

    rxnB_list = pd.DataFrame(np.array(rxnB_list).reshape(len(P_list), len(T_list))).T
    return rxnB_list

def Bmix(Aspen, Asp, T_list, P_list, n_list,NCOMP):
    aspen_tools.cd_properties(Asp)
    # Asp.set_focus()
    # Asp.child_window(title="MIX-1", control_type="TreeItem").click_input()
    B_mix_list = []

    Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\MIX-1\Input\LOWER\#1").Value = T_list[0]
    Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\MIX-1\Input\UPPER\#1").Value = T_list[-1]
    Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\MIX-1\Input\NPOINT\#1").Value = len(T_list) - 1
    x = Aspen.Application.Tree.FindNode("\Data\Properties\Analysis\MIX-1\Input\LIST\#0").Elements.Count
    for _ in range(x):
        Aspen.Application.Tree.FindNode("\Data\Properties\Analysis\MIX-1\Input\LIST\#0").Elements.RemoveRow(0,0)
    for _ in range(len(P_list)):
        Aspen.Application.Tree.FindNode("\Data\Properties\Analysis\MIX-1\Input\LIST\#0").Elements.InsertRow(0,0)
    

    for pi, P in enumerate(P_list):
        Aspen.Application.Tree.FindNode(f"\Data\Properties\Analysis\MIX-1\Input\LIST\#0\#{pi}").Value = P

    if isinstance(n_list[0],list):
        for ni, N in enumerate(n_list):
            for i in range(NCOMP):
                Aspen.Application.Tree.FindNode(
                    f"\Data\Properties\Analysis\MIX-1\Input\FLOW\COMP{i+1}"
                ).Value = N[i]

            ntot = sum(N)
            aspen_tools.run(Asp)
            
            B_mix = Aspen.Application.Tree.FindNode(
                r"\Data\Properties\Analysis\MIX-1\Output\Prop Data\PROPTAB"
            ).Elements.Item("TOTAL AVAILMX", str(ni + 1)).Value
            B_mix_list.append(round(B_mix * ntot, 7))
    else:
        for i in range(NCOMP):
            Aspen.Application.Tree.FindNode(f"\Data\Properties\Analysis\MIX-1\Input\FLOW\COMP{i+1}").Value = n_list[i]
        ntot = sum(n_list)
        aspen_tools.run(Asp)
        for i in range(len(T_list) * len(P_list)):
            B_mix = Aspen.Application.Tree.FindNode(r"\Data\Properties\Analysis\MIX-1\Output\Prop Data\PROPTAB").Elements.Item("TOTAL AVAILMX", str(i + 1)).Value
            B_mix_list.append(round(B_mix * ntot, 7))        

    B_mix_list = pd.DataFrame(np.array(B_mix_list).reshape(len(P_list), len(T_list))).T
    return B_mix_list

def Xeq(Aspen, Asp, T_list, P_list, n_in):
    Tl = T_list[0]
    Tu = T_list[-1]
    aspen_tools.cd_simulation(Asp)
    for i, a in enumerate(n_in):
        Aspen.Application.Tree.FindNode(f"\Data\Streams\FEED\Input\FLOW\MIXED\COMP{i+1}").Value = a
    x = Aspen.Application.Tree.FindNode("\Data\Model Analysis Tools\Sensitivity\XEQ\Input\LIST\#0").Elements.Count
    for _ in range(x):
        Aspen.Application.Tree.FindNode("\Data\Model Analysis Tools\Sensitivity\XEQ\Input\LIST\#0").Elements.RemoveRow(0,0)
    for _ in range(len(P_list)):
        Aspen.Application.Tree.FindNode("\Data\Model Analysis Tools\Sensitivity\XEQ\Input\LIST\#0").Elements.InsertRow(0,0)
    for pi, P in enumerate(P_list):
        Aspen.Application.Tree.FindNode(
            f"\Data\Model Analysis Tools\Sensitivity\XEQ\Input\LIST\#0\#{pi}"
        ).Value = P
    Aspen.Application.Tree.FindNode(
        r"\Data\Model Analysis Tools\Sensitivity\XEQ\Input\LOWER\#1"
    ).Value = Tl
    Aspen.Application.Tree.FindNode(
        r"\Data\Model Analysis Tools\Sensitivity\XEQ\Input\UPPER\#1"
    ).Value = Tu
    Aspen.Application.Tree.FindNode(
        r"\Data\Model Analysis Tools\Sensitivity\XEQ\Input\NPOINT\#1"
    ).Value = len(T_list)
    
    Asp.child_window(auto_id="MMTabItem_2", control_type="TabItem").select()
    try:
        aspen_tools.reset(Asp)
        aspen_tools.run_sim(Asp)
        n_out = []
        for j in range(len(P_list)):
            for i in range(len(T_list)):
                n_out_i = []
                for k in range(len(n_in)):
                    n_out_i.append(Aspen.Application.Tree.FindNode("\Data\Model Analysis Tools\Sensitivity\XEQ\Output\SENSVAR").Elements.Item(str(j * len(T_list) + i + 1), str(k+3)).Value)
                n_out.append(n_out_i)
    except AttributeError:
        aspen_tools.reset(Asp)
        aspen_tools.run_sim(Asp)
        n_out = []
        for j in range(len(P_list)):
            for i in range(len(T_list)):
                n_out_i = []
                for k in range(len(n_in)):
                    n_out_i.append(Aspen.Application.Tree.FindNode("\Data\Model Analysis Tools\Sensitivity\XEQ\Output\SENSVAR").Elements.Item(str(j * len(T_list) + i + 1), str(k+3)).Value)
                n_out.append(n_out_i)   
    return n_out

def calc_recycle(Aspen, Asp, n_in, n_out):
    n_inr = []
    n_outr = []
    n_out_after_recycle = []
    k_list=[]
    for i, a in enumerate(n_in):
        if a !=0:
            k_list.append(n_out[0][i]/a)
    k = min(k_list)
    k_index = k_list.index(k)
    rec_index = [i for i, val in enumerate(k_list) if val - k < 1e-6]
    string_indices = [str(i+1) for i in rec_index]
    indices_string = ", ".join(string_indices)
    print(f"COMP {indices_string} can be fully recycled")
    for i, n_out_i in enumerate(n_out):
        a = n_out_i[k_index]/(n_in[k_index]- n_out_i[k_index])
        n_inr_i = [(1+a)*x for x in n_in]
        n_outr_i = [(1+a)*x for x in n_out_i]
        n_inr.append(n_inr_i)
        n_outr.append(n_outr_i)
        n_out_after_recycle.append([ x - (a * y) for x, y in zip(n_outr_i, n_in)])
    return n_inr, n_outr, n_out_after_recycle
       

def calc_work(Aspen, Asp, Aspen_path,T_list, P_list, Tb_list, n_in, n_out,f,NCOMP,recycle=False):
    if recycle:
        n_inr, n_outr, n_out = calc_recycle(Aspen, Asp, n_in, n_out)
        
    print("Running n_in_Bp...")
    for attempt in range(3):
        try:
            n_in_Bp = Bpure(Aspen, Asp, T_list, P_list, Tb_list, n_in, NCOMP)
            break
        except pythoncom.com_error as e:
            
            print(f"Attempt {attempt + 1} failed: {e}")
            time.sleep(5)
            Aspen, Asp = aspen_tools.reconnect_aspen(Aspen_path)
        except Exception as e:
            print(f"Unexpected error: {e}")
            raise  
    if recycle:
        print("Running n_inr_Bp...")
        for attempt in range(3):
            try:
                n_inr_Bp = Bpure(Aspen, Asp, T_list, P_list,Tb_list, n_inr, NCOMP)
                break
            except pythoncom.com_error as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                time.sleep(5)
                Aspen, Asp = aspen_tools.reconnect_aspen(Aspen_path)
            except Exception as e:
                print(f"Unexpected error: {e}")
                raise        
        print("Running n_inr_Bm...")
        for attempt in range(3):
            try:
                n_inr_Bm = Bmix(Aspen, Asp, T_list, P_list, n_inr, NCOMP)
                break
            except pythoncom.com_error as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                time.sleep(5)
                Aspen, Asp = aspen_tools.reconnect_aspen(Aspen_path)
            except Exception as e:
                print(f"Unexpected error: {e}")
                raise
        print("Running n_outr_Bm...")
        for attempt in range(3):
            try:
                n_outr_Bm = Bmix(Aspen, Asp, T_list, P_list, n_outr, NCOMP)
                break
            except pythoncom.com_error as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                time.sleep(5)
                Aspen, Asp = aspen_tools.reconnect_aspen(Aspen_path)
            except Exception as e:
                print(f"Unexpected error: {e}")
                raise
        print("Running n_outr_Bp...")
        for attempt in range(3):
            try:
                n_outr_Bp = Bpure(Aspen, Asp, T_list, P_list, Tb_list,n_outr, NCOMP)
                break
            except pythoncom.com_error as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                time.sleep(5)
                Aspen, Asp = aspen_tools.reconnect_aspen(Aspen_path)
            except Exception as e:
                print(f"Unexpected error: {e}")
                raise
    else:
        print("Running n_in_Bm...")
        for attempt in range(3):
            try:
                n_in_Bm = Bmix(Aspen, Asp, T_list, P_list, n_in, NCOMP)
                break
            except pythoncom.com_error as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                time.sleep(5)
                Aspen, Asp = aspen_tools.reconnect_aspen(Aspen_path)
            except Exception as e:
                print(f"Unexpected error: {e}")
                raise  

        print("Running n_out_Bm...")
        for attempt in range(3):
            try:
                n_out_Bm = Bmix(Aspen, Asp, T_list, P_list, n_out, NCOMP)
                break
            except pythoncom.com_error as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                time.sleep(5)
                Aspen, Asp = aspen_tools.reconnect_aspen(Aspen_path)
            except Exception as e:
                print(f"Unexpected error: {e}")
                raise     
        
    print("Running n_out_Bp...")
    for attempt in range(3):
        try:
            n_out_Bp = Bpure(Aspen, Asp, T_list, P_list, Tb_list, n_out, NCOMP)
            break
        except pythoncom.com_error as e:
            print(f"Attempt {attempt + 1} failed: {e}")
            time.sleep(5)
            Aspen, Asp = aspen_tools.reconnect_aspen(Aspen_path)
        except Exception as e:
            print(f"Unexpected error: {e}")
            raise  
    print("Running n_out_Bp0...")
    for attempt in range(3):
        try:
            n_out_Bp0 =Bpure(Aspen, Asp, T_list, P_list, Tb_list, n_out, NCOMP, True)
            break
        except pythoncom.com_error as e:
            print(f"Attempt {attempt + 1} failed: {e}")
            time.sleep(5)
            Aspen, Asp = aspen_tools.reconnect_aspen(Aspen_path)
        except Exception as e:
            print(f"Unexpected error: {e}")
            raise  
    # n_in_Bp = Bpure(Aspen, Asp, T_list, P_list, Tb_list, n_in, NCOMP)
    # n_in_Bm = Bmix(Aspen, Asp, T_list, P_list, n_in, NCOMP)
    # n_out_Bm = Bmix(Aspen, Asp, T_list, P_list, n_out, NCOMP)
    # n_out_Bp = Bpure(Aspen, Asp, T_list, P_list, Tb_list, n_out, NCOMP)
    # n_out_Bp0 =Bpure(Aspen, Asp, T_list, P_list, Tb_list, n_out, NCOMP, True)
                                                    
    W_HC = n_in_Bp - n_in_Bp.iloc[0,0]
    if recycle:
        W_M = n_inr_Bm - n_inr_Bp
        W_R = n_outr_Bm - n_inr_Bm
        W_S = n_outr_Bp - n_outr_Bm
    else:
        W_M = n_in_Bm - n_in_Bp
        W_R = n_out_Bm - n_in_Bm
        W_S = n_out_Bp - n_out_Bm

    W_CD = n_out_Bp0 - n_out_Bp
    W_HC_sgen = W_HC.applymap(lambda x: x * 1/f if x >= 0 else x*f)
    W_M_sgen = W_M.applymap(lambda x: x * 1/f if x >= 0 else x*f)
    W_R_sgen = W_R.applymap(lambda x: x * 1/f if x >= 0 else x*f)
    W_S_sgen = W_S.applymap(lambda x: x * 1/f if x >= 0 else x*f)
    W_CD_sgen = W_CD.applymap(lambda x: x * 1/f if x >= 0 else x*f)

    n_out_1 = [item[0] for item in n_out]
    n_in_1 = n_in[0]  
    n_out_1_df = pd.DataFrame(np.array(n_out_1).reshape(len(P_list), len(T_list))).T

    W_total = W_HC_sgen + W_M_sgen + W_R_sgen + W_S_sgen + W_CD_sgen
    W_rev = W_HC + W_M + W_R + W_S + W_CD
    
    CF = (n_out_1_df - n_in_1 + 0.00321 * W_total) / n_in_1
    CF_delta = (n_out_1_df - n_in_1 + 0.00321 * W_total) / (n_in_1-n_out_1_df)

    return Aspen,Asp,W_total, W_rev, W_HC_sgen, W_M_sgen, W_R_sgen, W_S_sgen, W_CD_sgen, CF, CF_delta


def compute_Xeq_df(n_out, n_in, T_list, P_list, comp_index):
    NT = len(T_list)
    NP = len(P_list)
    n_in_comp = n_in[comp_index-1]
    Xeq = [[0 for _ in range(NP)] for _ in range(NT)]
    idx = 0
    for j in range(NP):           # column = P
        for i in range(NT):       # row = T
            n_out_comp = n_out[idx][comp_index-1]
            Xeq[i][j] = (n_in_comp - n_out_comp) / n_in_comp if n_in_comp != 0 else 0
            idx += 1

    df = pd.DataFrame(Xeq, index=T_list, columns=P_list)

    return df


def export_to_excel(filename, T_list, P_list,Tb_list, Xeq,W_total, W_rev, W_HC_sgen, W_M_sgen, W_R_sgen, W_S_sgen, W_CD_sgen,CF, CF_delta):
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        sheet_name = "Results"
        startcol = 0
        data_dict = {
            "W_total": W_total,
            "W_rev": W_rev,
            "W_HC_sgen": W_HC_sgen,
            "W_M_sgen": W_M_sgen,
            "W_R_sgen": W_R_sgen,
            "W_S_sgen": W_S_sgen,
            "W_CD_sgen": W_CD_sgen,
            "CF": CF,
            "CF_delta": CF_delta
        }

        for name, df in data_dict.items():
            df.columns = P_list
            df.index = T_list

            pd.DataFrame([[name]]).to_excel(writer, sheet_name=sheet_name, startcol=startcol, header=False, index=False)

            df.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=startcol)
            startcol += len(df.columns) + 4 

        Tb_df = pd.DataFrame(
            Tb_list,
            index=[f"{P} atm" for P in P_list],
            columns=[f"COMP{i+1}" for i in range(len(Tb_list[0]))]
        )
        pd.DataFrame([["Tb_list"]]).to_excel(writer, sheet_name=sheet_name, startcol=startcol, header=False, index=False)
        Tb_df.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=startcol)
        startcol += len(Tb_df.columns) + 4
        Xeq.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=startcol)
    wb = load_workbook(filename)
    ws = wb["Results"]

    font = Font(name="Calibri", bold=False)
    no_border = Border(
            left=Side(style=None),
            right=Side(style=None),
            top=Side(style=None),
            bottom=Side(style=None)
        )


    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.font = font
                cell.border = no_border

    wb.save(filename)